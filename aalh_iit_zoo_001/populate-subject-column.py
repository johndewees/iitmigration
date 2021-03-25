from openpyxl import load_workbook

filename = 'aalh_iit_zoo_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 115

iterationrow = 7
descol = 8
subcol = 9

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=descol).value
    for cell in row:
        if testvar.find('Park') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Parks. Photographs.; Animals. Photographs.'
        elif testvar.find('park') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Parks. Photographs.; Animals. Photographs.'
        elif testvar.find('Zoo') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Zoos. Photographs.; Animals. Photographs.'
        elif testvar.find('zoo') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Zoos. Photographs.; Animals. Photographs.'
        else:
            ws.cell(row=iterationrow, column=subcol).value = 'Animals. Photographs.'
    print(iterationrow,'|',ws.cell(row=iterationrow, column=subcol).value)
    iterationrow = iterationrow + 1
print('*****COMPLETED*****')
wb.save('aalh_iit_zoo_001.xlsx')