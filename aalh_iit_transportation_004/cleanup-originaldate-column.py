from openpyxl import load_workbook
import re

filename = 'aalh_iit_transportation_004.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 533

iterationrow = 7
targetcol = 15

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        cleandate = None
        approx = 'approximately '
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=targetcol).value = ''
            elif testvar.find('October 1984') != -1:
                continue
            elif testvar.startswith('c'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('C'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('a'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.endswith('?'):
                cleandate = testvar[:-1]
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate
            elif testvar.find('-') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find(',') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find('/') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            else :
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = cleandate[0]
            print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=targetcol).value)
        except:
            print(iterationrow,'|',testvar,'|','STATUS = PROBLEM')
    iterationrow = iterationrow + 1
wb.save('aalh_iit_transportation_004.xlsx')