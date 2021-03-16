from openpyxl import load_workbook

filename = 'aalh_iit_buildings_009.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 592

iterationrow = 7
targetcol = 13
titlecol = 2
desccol = 8

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=desccol).value
    for cell in row:
        if testvar.endswith('|'):
            desc = testvar[:-1]
            desc = desc.strip()
            ws.cell(row=iterationrow, column=desccol).value = desc
            print(iterationrow,'PIPE FOUND END')
        elif testvar.find(': |') != -1:
            desc2 = testvar.replace(': |',':')
            ws.cell(row=iterationrow, column=desccol).value = desc2
            print(iterationrow,'PIPE FOUND START')
        else:
            continue
    iterationrow = iterationrow + 1
print('***FINISHED SEARCHING FOR PIPES***')
wb.save("aalh_iit_buildings_009.xlsx")