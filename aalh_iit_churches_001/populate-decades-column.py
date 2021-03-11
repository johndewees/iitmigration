from openpyxl import load_workbook

filename = 'aalh_iit_churches_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 601

iterationrow = 7
targetcol = 15
decadescol = 14

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = str(ws.cell(row=iterationrow, column=targetcol).value)
        print(testvar)
        if testvar == None:
            ws.cell(row=iterationrow, column=decadescol).value = None
        elif testvar.find('180') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1800s'
        elif testvar.find('181') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1810s'
        elif testvar.find('182') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1820s'
        elif testvar.find('183') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1830s'
        elif testvar.find('184') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1840s'
        elif testvar.find('185') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1850s'
        elif testvar.find('186') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1860s'
        elif testvar.find('187') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1870s'
        elif testvar.find('188') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1880s'
        elif testvar.find('189') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1890s'
        elif testvar.find('190') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1900s'
        elif testvar.find('191') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1910s'
        elif testvar.find('192') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1920s'
        elif testvar.find('193') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1930s'
        elif testvar.find('194') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1940s'
        elif testvar.find('195') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1950s'
        elif testvar.find('196') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1960s'
        elif testvar.find('197') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1970s'
        elif testvar.find('198') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1980s'
        elif testvar.find('199') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '1990s'
        elif testvar.find('200') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '2000s'
        elif testvar.find('201') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '2010s'
        elif testvar.find('202') != -1:
            ws.cell(row=iterationrow, column=decadescol).value = '2020s'
        print(ws.cell(row=iterationrow, column=decadescol).value)
        iterationrow = iterationrow + 1
wb.save('aalh_iit_churches_001.xlsx')