from openpyxl import load_workbook
import re

filename = 'aalh_iit_parksnature_002.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 144

iterationrow = 7
targetcol = 2
locationcol = 13
datecol = 15
isodatecol = 16

commaspace = ', '
space = ' '
openbracket = '['
closebrakcet = ']'

states = ['Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut','Delaware','Florida','Georgia','Hawaii','Idaho','Illinois','Indiana','Iowa','Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts','Michigan','Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada','New Hampshire','New Jersey','New Mexico','New York','North Carolina','North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island','South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont','Virginia','Washington','West Virginia','Wisconsin','Wyoming']

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        locationvar = ws.cell(row=iterationrow, column=locationcol).value
        if locationvar == None:
            locationvarfinal = ''
            #print(locationvarfinal)
        elif locationvar in states:
            locationvarfinal = commaspace + locationvar
            #print(locationvarfinal)
        else :
            locationvar1 = locationvar.split(';')
            locationvar2 = locationvar1[0]
            locationvar3 = locationvar2.split('(')
            locationvar4a = locationvar3[0]
            locationvar4b = locationvar3[1]
            locationvar5a = locationvar4a.strip()
            locationvar5b = locationvar4b[:-1]
            locationvarfinal = commaspace + locationvar5a + commaspace + locationvar5b
            #print(locationvarfinal)
    for cell in row:
        datevar = ws.cell(row=iterationrow, column=datecol).value
        datevar2 = ws.cell(row=iterationrow, column=isodatecol).value
        if datevar == None:
            datevarfinal = ' [date unknown]'
            #print(datevarfinal)
        elif datevar2 == None:
            datevarfinal = ' [date unknown]'
        else:
            datevar = str(datevar)
            if datevar.startswith("approximately"):
                if datevar.find('-') != -1:
                    datevarregex = re.findall('\d\d\d\d', datevar)
                    datevarfinal = space + openbracket + 'approximately ' + datevarregex[0] + closebrakcet
                else:
                    datevarfinal = space + openbracket + datevar + closebrakcet
                #print(datevarfinal)
            elif datevar.find('-') != -1:
                datevarregex = re.findall('\d\d\d\d', datevar)
                datevarfinal = commaspace + datevarregex[0]
                #print(datevarfinal)
            else:
                datevarfinal = commaspace + datevar
                #print(datevarfinal)
    for cell in row:
        title = str(ws.cell(row=iterationrow, column=targetcol).value)
        titlefinal = title + locationvarfinal + datevarfinal
        ws.cell(row=iterationrow, column=targetcol).value = titlefinal
        print(iterationrow,'|',titlefinal)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_parksnature_002.xlsx')