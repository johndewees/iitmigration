from openpyxl import load_workbook
import re

filename = 'aalh_iit_industry_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 472

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = None
            elif testvar.find('1921-1923') != -1:
                isovalue = '1921; 1922; 1923'
                ws.cell(row=iterationrow, column=targetcol).value = '1921-1923'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('September, 1921') != -1:
                isovalue = '1921-09'
                ws.cell(row=iterationrow, column=targetcol).value = 'September 1921'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('March, 1908') != -1:
                isovalue = '1908-03'
                ws.cell(row=iterationrow, column=targetcol).value = 'March 1908'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('January, 1924') != -1:
                isovalue = '1924-01'
                ws.cell(row=iterationrow, column=targetcol).value = 'January 1924'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('-') != -1:
                isovalue = testvar
                if isovalue.endswith('?'):
                    isovalue = isovalue[:-1]
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                print(iterationrow,'|',testvar,'|','STATUS = AMERICAN DATE')
            elif testvar.find('/') != -1:
                print(iterationrow,'|',testvar,'|','STATUS = AMERICAN DATE')
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            print(iterationrow,'|',testvar,'|','STATUS = PROBLEM')
        print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=isostandardcol).value)
    for cell in row:
        testvar2 = ws.cell(row=iterationrow, column=isostandardcol).value
        if testvar2 == None:
            continue
        elif testvar2.endswith('?'):
            testvar3 = testvar2[:-1]
            ws.cell(row=iterationrow, column=isostandardcol).value = testvar3
            print(iterationrow,'| TRIMMED ?')
    iterationrow = iterationrow + 1
wb.save('aalh_iit_industry_001.xlsx')