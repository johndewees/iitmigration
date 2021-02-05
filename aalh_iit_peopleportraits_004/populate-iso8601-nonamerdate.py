from openpyxl import load_workbook
import re

filename = 'aalh_iit_peopleportraits_004.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 501

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = None
            elif testvar.find('1945-46') != -1:
                isovalue = '1945; 1946'
                ws.cell(row=iterationrow, column=targetcol).value = '1945-1946'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('-') != -1:
                isovalue = testvar
                if isovalue.endswith('?'):
                    isovalue = isovalue[:-1]
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                print('STATUS = AMERICAN DATE')
            elif testvar.find('/') != -1:
                print('STATUS = AMERICAN DATE')
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            print('STATUS = PROBLEM')
        print(ws.cell(row=iterationrow, column=isostandardcol).value)
    for cell in row:
        testvar2 = ws.cell(row=iterationrow, column=isostandardcol).value
        if testvar2 == None:
            continue
        elif testvar2.endswith('?'):
            testvar3 = testvar2[:-1]
            ws.cell(row=iterationrow, column=isostandardcol).value = testvar3
    iterationrow = iterationrow + 1
wb.save('aalh_iit_peopleportraits_004.xlsx')