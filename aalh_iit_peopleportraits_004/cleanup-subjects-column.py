from openpyxl import load_workbook

filename = 'aalh_iit_peopleportraits_004.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 501

iterationrow = 7
targetcol = 9
placecol = 13

subjectholland = 'Images in time photographic collection. (Toledo Lucas County Public Library); Holland (Ohio). History. Photographs.'
subjectwaterville = 'Images in time photographic collection. (Toledo Lucas County Public Library); Waterville (Ohio). History. Photographs.'
subjectoregon = 'Images in time photographic collection. (Toledo Lucas County Public Library); Oregon (Ohio). History. Photographs.'
subjectmaumee = 'Images in time photographic collection. (Toledo Lucas County Public Library); Maumee (Ohio). History. Photographs.'
subjectsylvania = 'Images in time photographic collection. (Toledo Lucas County Public Library); Sylvania (Ohio). History. Photographs.'
subjecttoledo = 'Images in time photographic collection. (Toledo Lucas County Public Library); Toledo (Ohio). History. Photographs.'
subjectnonlucascounty = 'Images in time photographic collection. (Toledo Lucas County Public Library)'
semicolonspace = '; '

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        placevar = ws.cell(row=iterationrow, column=placecol).value
        if testvar == None:
            if placevar == None:
                ws.cell(row=iterationrow, column=targetcol).value = subjectnonlucascounty
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Toledo (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjecttoledo
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Sylvania (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjectsylvania
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Maumee (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjectmaumee
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Oregon (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjectoregon
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Waterville (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjectwaterville
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Holland (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = subjectholland
                print(ws.cell(row=iterationrow, column=targetcol).value)
            else:
                ws.cell(row=iterationrow, column=targetcol).value = subjectnonlucascounty
                print(ws.cell(row=iterationrow, column=targetcol).value)
        else:
            if placevar == None:    
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectnonlucascounty
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Toledo') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjecttoledo
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Sylvania (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectsylvania
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Maumee (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectmaumee
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Oregon (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectoregon
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Waterville (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectwaterville
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif placevar.find('Holland (Ohio)') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectholland
                print(ws.cell(row=iterationrow, column=targetcol).value)
            else:
                ws.cell(row=iterationrow, column=targetcol).value = testvar + semicolonspace + subjectnonlucascounty
                print(ws.cell(row=iterationrow, column=targetcol).value)
        iterationrow = iterationrow + 1
wb.save('aalh_iit_peopleportraits_004.xlsx')