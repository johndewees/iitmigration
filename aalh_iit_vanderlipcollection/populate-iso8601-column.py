from openpyxl import load_workbook
import re

#enter the source filename for the Excel worksheet in this variable
filename = 'aalh_iit_vanderlipcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

#variables define the array of rows and columns
minimumcol = 15
maximumcol = 15
minimumrow = 494
maximumrow = 899
#the iteration variable used in the function
iterationrow = 494
#the column from which the data is being read
targetcol = 15
#the column that will be populated with new data
isostandardcol = 16

for row in ws.iter_rows(min_row=iterationrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(cell.value, end=" ")
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        isovalue = None
        if testvar == None:
            ws.cell(row=iterationrow, column=isostandardcol).value = ''
        else :
            isovalue = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        print(ws.cell(row=iterationrow, column=isostandardcol).value)
        iterationrow = iterationrow + 1
#establish target file to be saved here
wb.save("aalh_iit_vanderlipcollection.xlsx")