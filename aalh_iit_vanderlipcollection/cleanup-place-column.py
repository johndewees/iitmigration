from openpyxl import load_workbook

#enter the source filename for the Excel worksheet in this variable
filename = 'aalh_iit_vanderlipcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

#variables define the array of rows and columns
minimumcol = 8
maximumcol = 8
minimumrow = 530
maximumrow = 899

iterationrow = 530
targetcol = 13

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=minimumcol).value
        if testvar.find('Philadelphia') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Philadelphia (Pennsylvania); Philadelphia County (Pennsylvania)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Cleveland') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Cleveland (Ohio); Cuyahoga County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Upper Sandusky') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Upper Sandusky (Ohio); Wyandot County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Sandusky') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Sandusky (Ohio); Erie County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Mt. Vernon') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Mt. Vernon (Ohio); Knox County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Detroit') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Detroit (Michigan); Wayne County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Sylvania') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Sylvania (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Oregon') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Oregon (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Waterville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Waterville (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Perrysburg') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Perrysburg (Ohio); Wood County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        else :
            print('No changes needed')
        iterationrow = iterationrow + 1

#establish target file to be saved here
wb.save("aalh_iit_vanderlipcollection.xlsx")