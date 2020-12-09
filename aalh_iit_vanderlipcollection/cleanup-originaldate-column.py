from openpyxl import load_workbook
import re

#enter the source filename for the Excel file in this variable
filename = 'aalh_iit_vanderlipcollection.xlsx'
#loads the entire Excel workbook
wb = load_workbook(filename)
#loads the specific sheet in the workbook to be manipulated
ws = wb['Metadata Template']

#variables define the array of rows and columns
minimumcol = 15
maximumcol = 15
minimumrow = 494
maximumrow = 899
#the iteration variable used in the function
iterationrow = 494
#the column from which the data is being read
targetcol = 15

for row in ws.iter_rows(min_row=iterationrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(cell.value, end=" ")
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        cleandate = None
        approx = 'approximately '
        if testvar == None:
            ws.cell(row=iterationrow, column=targetcol).value = ''
        elif testvar.endswith('?'):
            cleandate = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
        elif testvar.startswith('c'):
            cleandate = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
        elif testvar.startswith('C'):
            cleandate = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
        elif testvar.startswith('a'):
            cleandate = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
        else :
            cleandate = re.findall('\d\d\d\d', testvar)
            ws.cell(row=iterationrow, column=targetcol).value = cleandate[0]
        print(ws.cell(row=iterationrow, column=targetcol).value)
        iterationrow = iterationrow + 1
        
#establish target file to be saved here
wb.save("aalh_iit_vanderlipcollection.xlsx")