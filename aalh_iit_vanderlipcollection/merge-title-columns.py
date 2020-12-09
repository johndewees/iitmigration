from openpyxl import load_workbook

#enter the source filename for the Excel worksheet in this variable
filename = 'aalh_iit_vanderlipcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

#variables define the array of rows and columns
minimumcol = 2
maximumcol = 2
minimumrow = 494
maximumrow = 898

iterationrow = 494
targetcol = 2
locationcol = 13
datecol = 15

commaspace = ', '
space = ' '
openbracket = '['
closebrakcet = ']'

states = ['Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut','Delaware','Florida','Georgia','Hawaii','Idaho','Illinois','Indiana','Iowa','Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts','Michigan','Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada','New Hampshire','New Jersey','New Mexico','New York','North Carolina','North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island','South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont','Virginia','Washington','West Virginia','Wisconsin','Wyoming']

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        locationvar = ws.cell(row=iterationrow, column=locationcol).value
        if locationvar == None:
            print('No location')
        elif locationvar in states:
            locationvarfinal = locationvar
            print(locationvarfinal)
        else :
            #print(locationvar)
            locationvar1 = locationvar.split(';')
            #print(locationvar1)
            locationvar2 = locationvar1[0]
            #print(locationvar2)
            locationvar3 = locationvar2.split('(')
            #print(locationvar3)
            locationvar4a = locationvar3[0]
            locationvar4b = locationvar3[1]
            #print(locationvar4a)
            #print(locationvar4b)
            locationvar5a = locationvar4a.strip()
            locationvar5b = locationvar4b[:-1]
            #print(locationvar5a)
            #print(locationvar5b)
            locationvarfinal = locationvar5a + commaspace + locationvar5b
            print(locationvarfinal)
    for cell in row:
        datevar = str(ws.cell(row=iterationrow, column=datecol).value)
        datevar = datevar
        if datevar == None:
            print('No date')
        elif datevar.startswith("approximately"):
            datevarfinal = openbracket + datevar + closebrakcet
            print(datevarfinal)
        else:
            datevarfinal = datevar
            print(datevarfinal)
    for cell in row:
        title = str(ws.cell(row=iterationrow, column=targetcol).value)
        if datevarfinal.startswith('[approximately'):
            titlefinal = title + commaspace + locationvarfinal + space + datevarfinal
        else:
            titlefinal = title + commaspace + locationvarfinal + commaspace + datevarfinal
        ws.cell(row=iterationrow, column=targetcol).value = titlefinal
        print(titlefinal)
    iterationrow = iterationrow + 1

#establish target file to be saved here
wb.save("aalh_iit_vanderlipcollection.xlsx")