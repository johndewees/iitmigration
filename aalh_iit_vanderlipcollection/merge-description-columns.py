from openpyxl import load_workbook

#enter the filename for the Excel spreadsheet for the final metadata that will be uploaded to CDM
filename = 'testing.xlsx'
#loads the workbooks associated with the two files above, 1 and 2 corresponding to the filenames
wb = load_workbook(filename)
#loads the worksheets associated with the two files above, 1 and 2 corresponding to the filenames
ws1 = wb['Metadata Template']
ws2 = wb['keywords']

#variables define the array of rows and columns in ws1 to be iterated over
minimumcol = 8
maximumcol = 8
minimumrow = 494
maximumrow = 899
#the iteration variables used in the function
iterationrow1 = 494
iterationrow2 = 1
#the column from which the data is being read
targetcol = 8
space = ' '

for row in ws1.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar1 = ws1.cell(row=iterationrow1, column=targetcol).value
        print(testvar1)
        testvar2 = ws2.cell(row=iterationrow2, column=2).value
        print(testvar2)
        ws1.cell(row=iterationrow1, column=targetcol).value = testvar1 + space + testvar2
        print(ws1.cell(row=iterationrow1, column=targetcol).value)

        iterationrow1 = iterationrow1 + 1
        iterationrow2 = iterationrow2 + 1

#establish target file to be saved here
wb.save("testing-iterated.xlsx")