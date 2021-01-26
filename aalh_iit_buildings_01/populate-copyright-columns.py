from openpyxl import load_workbook
import re

filename = 'aalh_iit_buildings_01.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 504

iterationrow = 7
targetcol = 15
rightsuricol = 26
rightsdesccol = 27
nocruri = 'http://rightsstatements.org/vocab/NoC-US/1.0/'
nocrdesc = 'No copyright - United States'
incruri = 'http://rightsstatements.org/vocab/InC/1.0/'
incrdesc = 'In copyright'

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        try:
            if testvar == None:
                print('NO DATE TO WORK WITH')
            else:
                testyearregex = re.findall('\d\d\d\d', testvar)
                testyear1 = testyearregex[0]
                testyear2 = testyear1.strip()
                testyear = int(testyear2)
                if testyear < 1926:
                    ws.cell(row=iterationrow, column=rightsuricol).value = nocruri
                    ws.cell(row=iterationrow, column=rightsdesccol).value = nocrdesc
                    print('[',ws.cell(row=iterationrow, column=rightsuricol).value,']','[',ws.cell(row=iterationrow, column=rightsdesccol).value,']')
                elif testyear >= 1926:
                    ws.cell(row=iterationrow, column=rightsuricol).value = incruri
                    ws.cell(row=iterationrow, column=rightsdesccol).value = incrdesc
                    print('[',ws.cell(row=iterationrow, column=rightsuricol).value,']','[',ws.cell(row=iterationrow, column=rightsdesccol).value,']')
        except:
            print('STATUS = PROBLEM')
        iterationrow = iterationrow + 1
wb.save('aalh_iit_buildings_01.xlsx')