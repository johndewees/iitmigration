from openpyxl import load_workbook
import urllib.request, urllib.error

filename = 'aalh_iit_buildings_01.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 43
maximumcol = 43
minimumrow = 7
maximumrow = 504

iterationrow = 7
targetcol = 48

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        url = ws.cell(row=iterationrow, column=targetcol).value
        try:
            conn = urllib.request.urlopen(url)
            print('URL OPEN = SUCCESSFUL')
        except urllib.error.HTTPError as e:
            print(iterationrow)
            print('HTTPError: {}'.format(e.code))
        except urllib.error.URLError as e:
            print(iterationrow)
            print('URLError: {}'.format(e.reason))
    iterationrow = iterationrow + 1