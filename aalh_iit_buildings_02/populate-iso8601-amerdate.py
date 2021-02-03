from openpyxl import load_workbook
import re

filename = 'aalh_iit_buildings_02.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 506

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    print(ws.cell(row=iterationrow, column=targetcol).value)
    try:
        for cell in row:
            ameryear = None
            yearraw = ws.cell(row=iterationrow, column=targetcol).value
            if yearraw.find(',') != -1:
                ameryearre = re.findall('\d\d\d\d', yearraw)
                ameryear = ameryearre[0]
                print(ameryear)
            else:
                print('Not an American formatted date (year)')
        for cell in row:
            amermon = None
            monraw = ws.cell(row=iterationrow, column=targetcol).value
            if monraw.find(',') != -1:
                if monraw.find('Jan') != -1:
                    amermon = '01'
                elif monraw.find('jan') != -1:
                    amermon = '01'
                elif monraw.find('Feb') != -1:
                    amermon = '02'
                elif monraw.find('feb') != -1:
                    amermon = '02'
                elif monraw.find('Mar') != -1:
                    amermon = '03'
                elif monraw.find('mar') != -1:
                    amermon = '03'
                elif monraw.find('Apr') != -1:
                    amermon = '04'
                elif monraw.find('apr') != -1:
                    amermon = '04'
                elif monraw.find('May') != -1:
                    amermon = '05'
                elif monraw.find('may') != -1:
                    amermon = '05'
                elif monraw.find('Jun') != -1:
                    amermon = '06'
                elif monraw.find('jun') != -1:
                    amermon = '06'
                elif monraw.find('Jul') != -1:
                    amermon = '07'
                elif monraw.find('jul') != -1:
                    amermon = '07'
                elif monraw.find('aUG') != -1:
                    amermon = '08'
                elif monraw.find('aug') != -1:
                    amermon = '08'
                elif monraw.find('Sep') != -1:
                    amermon = '09'
                elif monraw.find('sep') != -1:
                    amermon = '09'
                elif monraw.find('Oct') != -1:
                    amermon = '10'
                elif monraw.find('oct') != -1:
                    amermon = '10'
                elif monraw.find('Nov') != -1:
                    amermon = '11'
                elif monraw.find('nov') != -1:
                    amermon = '11'
                elif monraw.find('Dec') != -1:
                    amermon = '12'
                elif monraw.find('dec') != -1:
                    amermon = '12'
                print(amermon)
            else:
                print('Not an American formatted date (month)')
        for cell in row:
            amerday = None
            dayraw = ws.cell(row=iterationrow, column=targetcol).value
            if dayraw.find(',') != -1:
                daypart1 = dayraw.split(',')
                daypart2 = daypart1[0]
                daypart3 = daypart2.split()
                daypart4 = daypart3[1]
                if daypart4.startswith('1'):
                    amerday = daypart4
                elif daypart4.startswith('2'):
                    amerday = daypart4
                elif daypart4.startswith('3'):
                    amerday = daypart4
                else:
                    amerday = '0' + daypart4
                print(amerday)
            else:
                print('Not an American formatted date (day)')
        for cell in row:
            testvar = ws.cell(row=iterationrow, column=targetcol).value
            if testvar.find('/') != -1:
                testvarlist = testvar.split('/')
                testvaryear = testvarlist[2]
                testvaryear = testvaryear.strip()
                testvarmonth = testvarlist[0]
                testvarmonth = testvarmonth.strip()
                testvarmonth = int(testvarmonth)
                if testvarmonth < 10:
                    testvarmonth = str(testvarmonth)
                    testvarmonth = '0' + testvarmonth
                else:
                    testvarmonth = str(testvarmonth)
                testvarday = testvarlist[1]
                testvarday = testvarday.strip()
                testvarday = int(testvarday)
                if testvarday < 10:
                    testvarday = str(testvarday)
                    testvarday = '0' + testvarday
                else:
                    testvarday = str(testvarday)
                isodate = testvaryear + '-' + testvarmonth + '-' + testvarday
                ws.cell(row=iterationrow, column=targetcol).value = isodate
                print(isodate)
            else:
                print ('Not a date formatted with a slash')
        for cell in row:
            if ameryear == None:
                print('Not an American formatted date at all')
            else:
                amerdatetrans = ameryear + '-' + amermon + '-' + amerday
                ws.cell(row=iterationrow, column=isostandardcol).value = amerdatetrans
                print(amerdatetrans)
    except:
        print('Not an American formatted date at all')
    iterationrow = iterationrow + 1
#wb.save('aalh_iit_buildings_02.xlsx')