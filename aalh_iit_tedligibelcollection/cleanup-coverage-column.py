from openpyxl import load_workbook

filename = 'aalh_iit_tedligibelcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 10
maximumcol = 10
minimumrow = 7
maximumrow = 2651

iterationrow = 7


for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=minimumcol).value
        if testvar == None:
            print('Nothing to work with')
        elif testvar.find('St.') != -1:
            testvar.replace("St.", "Street")
            print(testvar)
            ws.cell(row=iterationrow, column=minimumcol).value = testvar
            print(ws.cell(row=iterationrow, column=minimumcol).value)
        else:
            print('Nothing needs changed')
    iterationrow = iterationrow + 1
#wb.save('aalh_iit_tedligibelcollection.xlsx')