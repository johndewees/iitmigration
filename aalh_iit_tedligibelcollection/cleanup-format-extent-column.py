from openpyxl import load_workbook

filename = 'aalh_iit_tedligibelcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 47
maximumcol = 47
minimumrow = 7
maximumrow = 5141

iterationrow = 7
targetcol = 47
formatcol = 33
extentcol = 34

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        formatvar = ws.cell(row=iterationrow, column=extentcol).value
        print(formatvar)
        if formatvar == None:
            print('FIELD = NONE')
        elif formatvar == 'H: 2 in, W: 2 in':
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Slides'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        else :
            continue
    iterationrow = iterationrow + 1
wb.save('aalh_iit_tedligibelcollection.xlsx')