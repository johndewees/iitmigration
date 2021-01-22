from openpyxl import load_workbook

filename = 'aalh_iit_tedligibelcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 5141

iterationrow = 7

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        iitdescription = ws.cell(row=iterationrow, column=minimumcol).value
        if iitdescription == None:
            print('No description')
        elif iitdescription.endswith(','):
            description1 = iitdescription
            description2 = description1[:-1]
            description3 = description2 + '.'
            ws.cell(row=iterationrow, column=minimumcol).value = description3
            print(ws.cell(row=iterationrow, column=minimumcol).value)
        iterationrow = iterationrow + 1
    wb.save('aalh_iit_tedligibelcollection.xlsx')