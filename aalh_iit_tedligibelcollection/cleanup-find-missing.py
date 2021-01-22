from openpyxl import load_workbook
import urllib.request

filename = 'aalh_iit_tedligibelcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

textfile = '1-21-2021.txt'
texthandle = open(textfile)
textdata = texthandle.readlines()

minimumcol = 2
maximumcol = 45
minimumrow = 7
maximumrow = 2638
targetcol = 43
iterationrow = 7

uploaded = list()
notuploaded = list()

for textfilename1 in textdata:
    textfilename2 = textfilename1.strip()
    result = textfilename2.endswith("jpg")
    if result == True:
        uploaded.append(textfilename2)

for row in ws.iter_rows(min_row=minimumrow, min_col=targetcol, max_row=maximumrow, max_col=targetcol):
    for cell in row:
        excelfilename = ws.cell(row=iterationrow, column=targetcol).value
        if excelfilename in uploaded:
            continue
        else:
            notuploaded.append(iterationrow)
            notuploaded.append(excelfilename)
            downloadurl = ws.cell(row=iterationrow, column=48).value
            downloadfilename = ws.cell(row=iterationrow, column=43).value
            download = urllib.request.urlretrieve(downloadurl, downloadfilename)
    iterationrow = iterationrow + 1

for item in notuploaded:
    print(item)