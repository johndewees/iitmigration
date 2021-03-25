from openpyxl import load_workbook

filename = 'aalh_iit_jeep_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 47
maximumcol = 47
minimumrow = 7
maximumrow = 396

iterationrow = 7
targetcol = 47
formatcol = 33
extentcol = 34

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        extentvar = ws.cell(row=iterationrow, column=targetcol).value
        if extentvar == None:
            continue
        elif extentvar != None:
            if extentvar.find(' x ') != -1:
                extentvar1 = extentvar.split(';')
                for item in extentvar1:
                    if item.find(' x ') != -1:
                        extentvar2 = item
                        extentvar3 = extentvar2.split('x')
                        extentvar3a = extentvar3[0]
                        extentvar3b = extentvar3[1]
                        extentvar4a = extentvar3a.strip()
                        extentvar4b = extentvar3b.strip()
                        extentvarfinal = extentvar4a + ' in x ' + extentvar4b + ' in'
                        ws.cell(row=iterationrow, column=extentcol).value = extentvarfinal
                        print(iterationrow,'|',ws.cell(row=iterationrow, column=extentcol).value)
    iterationrow = iterationrow + 1
#wb.save('aalh_iit_jeep_001.xlsx')