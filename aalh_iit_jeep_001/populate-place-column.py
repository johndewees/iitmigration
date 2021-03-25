from openpyxl import load_workbook

filename = 'aalh_iit_jeep_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 396

iterationrow = 7
desccol = 8
targetcol = 13
rawcovcol = 49
covcol = 10
titlecol = 2

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    placefield = ws.cell(row=iterationrow, column=targetcol).value
    testvar = ws.cell(row=iterationrow, column=desccol).value
    for cell in row:
        print(iterationrow)
        try:
            if placefield != None: 
                print('FIELD ALREADY POPULATED')
            elif placefield == None:
                if testvar.find('Philadelphia') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Philadelphia (Pennsylvania); Philadelphia County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('McKean') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'McKean (Pennsylvania); Erie County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Elk Creek Township') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Elk Creek Township (Pennsylvania); Erie County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Cleveland') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Cleveland (Ohio); Cuyahoga County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Elmore') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Elmore (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bucyrus') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bucyrus (Ohio); Crawford County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Columbus') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Columbus (Ohio); Franklin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Okolona') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Okolona (Ohio); Henry County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Holgate') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Holgate (Ohio); Henry County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Napoleon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Napoleon (Ohio); Henry County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Upper Sandusky') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Upper Sandusky (Ohio); Wyandot County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sandusky') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Sandusky (Ohio); Erie County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Castalia') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Castalia (Ohio); Erie County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Mt. Vernon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Mt. Vernon (Ohio); Knox County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Hometown') != -1:
                    if testvar.find('Pennsylvania') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Hometown (Pennsylvania); Schuylkill County (Pennsylvania)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Detroit') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Detroit (Michigan); Wayne County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Flint') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Flint (Michigan); Genesee County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sylvania') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Sylvania (Ohio); Lucas County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oregon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oregon (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Holland') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Holland (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Holland (Ohio); Lucas County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('New Bremen') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'New Bremen (Ohio); Auglaize County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Brooklyn') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Brooklyn (Michigan); Jackson County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oshtemo') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oshtemo (Michigan); Kalamazoo County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Dowagiac') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Dowagiac (Michigan); Cass County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Knoxville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Knoxville (Tennessee); Knox County (Tennessee)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ada, Ohio') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ada (Ohio); Hardin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Antwerp') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Antwerp (Ohio); Paulding County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Richfield Center, Ohio') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Richfield Center (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Waterville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Waterville (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Richfield Township') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Richfield Township (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bowling Green') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bowling Green (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Stony Ridge') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Stony Ridge (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Rossford') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Rossford (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Walbridge') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Walbridge (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Perrysburg') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Perrysburg (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Cincinnati') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Cincinnati (Ohio); Hamilton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Hamilton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Hamilton (Ohio); Butler County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ottokee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ottokee (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Archbold') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Archbold (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Marblehead') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Marblehead (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Findlay') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Findlay (Ohio); Hancock County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delta') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delta (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Genoa') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Genoa (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Grand Rapids') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Ohio); Wood County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Michigan); Kent County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Niagara Falls') != -1:
                    if testvar.find('New York') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Niagara Falls (New York); Niagara County (New York)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Canada') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Niagara Falls (Canada)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Buffalo') != -1:
                    if testvar.find('New York') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Buffalo (New York); Erie County (New York)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delaware') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Delaware (Ohio); Delaware County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Erie') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Erie Township (Michigan); Monroe County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Pennsylvania') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Erie (Pennsylvania); Erie County (Pennsylvania)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Galloway') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Galloway (Ohio); Franklin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bryan') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bryan (Ohio); Williams County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pemberville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pemberville (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bradner') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bradner (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Port Clinton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Port Clinton (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Middle Bass') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Middle Bass Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('South Bass') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'South Bass Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Gibraltar Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Gibraltar Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find("Kelley's Island") != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Kelleys Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Kelleys Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Kelleys Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Catawba') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Catawba Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Monroeville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Monroeville (Ohio); Huron County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Dundee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Dundee (Michigan); Monroe County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Ignance') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'St. Ignance (Michigan); Mackinac County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Mackinac Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Mackinac Island (Michigan); Mackinac County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Monroe') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Monroe (Michigan); Monroe County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Texas') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Texas (Ohio); Henry County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Florida') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Florida (Ohio); Henry County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fremont') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fremont (Ohio); Sandusky County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Woodville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Woodville (Ohio); Sandusky County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Hessville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Hessville (Ohio); Sandusky County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fayette') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fayette (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Springfield') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Springfield (Ohio); Clark County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Illinois') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Springfield (Illinois); Sangamon County (Illinois)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ann Arbor') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ann Arbor (Michigan); Washtenaw County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Tiffin') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Tiffin (Ohio); Seneca County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lockington') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lockington (Ohio); Shelby County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delphos') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delphos (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sault St. Marie') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Sault St. Marie (Michigan), Chippewa County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sault Ste. Marie') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Sault St. Marie (Michigan), Chippewa County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Defiance') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Defiance (Ohio); Defiance County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Jewell') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Jewell (Ohio); Defiance County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Spencerville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Spencerville (Ohio); Allen County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Louisville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Louisville (Kentucky); Jefferson County (Kentucky)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Shreveport') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Shreveport (Louisiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Put-in-Bay') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Put-in-Bay Township (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Paul') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Saint Paul') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Providence') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Providence Township (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ottawa Hills') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ottawa Hills (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bono') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bono (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Winona') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Winona (Minnesota); Winona County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Duluth') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Duluth (Minnesota); St. Louis County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Galveston') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Galveston (Texas); Galveston County (Texas)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Camden') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Camden (New Jersey); Camden County (New Jersey)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Glassboro') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Glassboro (New Jersey); Gloucester County (New Jersey)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Swanton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Swanton (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Highland Park') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Highland Park (Michigan); Wayne County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('New Hudson') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'New Hudson (Michigan); Oakland County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Clay Center') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Clay Center (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Greenville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Greenville (Ohio); Darke County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lakeside') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lakeside (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Wauseon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Wauseon (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oak Harbor') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oak Harbor (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Stryker') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Stryker (Ohio); Williams County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Akron') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Akron (Ohio); Summit County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Johnsons Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Johnson's Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Green Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Green Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sugar Island') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = "Sugar Island (Ohio); Ottawa County (Ohio)"
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = "Sugar Island (Michigan); Chippewa County (Michigan)"
                        print(ws.cell(row=iterationrow, column=targetcol).value)    
                elif testvar.find("Johnson's Island") != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Johnson's Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Riga') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Riga Township (Michigan); Lenawee County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Grandville') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grandville (Michigan); Kent County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Decatur') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Decatur (Illinois); Macon County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Versailles') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Versailles (Illinois); Brown County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Naples') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Naples (Illinois); Scott County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Quincy') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Quincy (Illinois); Adams County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Meredosia') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Meredosia (Illinois); Morgan County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Adrian') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Adrian (Michigan); Lenawee County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Harrison') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Harrison (Michigan); Clare County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oberlin') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oberlin (Ohio); Lorain County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Milwaukee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Milwaukee (Wisconsin); Milwaukee County (Wisconsin)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lake Geneva, Wisconsin') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lake Geneva (Wisconsin); Walworth County (Wisconsin)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ballast Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ballast Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Marion') != -1:
                    if testvar.find('Indiana') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Marion (Indiana); Grant County (Indiana)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Marion (Ohio); Marion County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Peru') != -1:
                    if testvar.find('Indiana') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Peru (Indiana); Miami County (Indiana)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fort Wayne') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fort Wayne (Indiana); Allen County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('New Haven') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'New Haven (Indiana); Allen County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Huntington') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Huntington (Indiana); Huntington County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Indianapolis') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Indianapolis (Indiana); Marion County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Terra Haute') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Terra Haute (Indiana); Vigo County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Williamsport') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Williamsport (Indiana); Warren County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Wabash') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Wabash (Indiana); Wabash County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lagro') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lagro (Indiana); Wabash County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Mt. Sterling') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Mount Sterling (Indiana); Switzerland County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lafayette') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lafayette (Indiana); Tippecanoe County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Colburn') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Colburn (Indiana); Tippecanoe County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Logansport') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Logansport (Indiana); Cass County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delphi') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delphi (Indiana); Carroll County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Attica') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Attica (Indiana); Fountain County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Baltimore') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Baltimore (Maryland)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Toledo') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Toledo (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Jacksonville') != -1 :
                    if testvar.find('Florida') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Jacksonville (Florida); Duval County (Florida)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Illinois') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Jacksonville (Illinois); Morgan County (Illinois)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pittsburgh') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pittsburgh (Pennsylvania); Allegheny County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Easttown') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Easttown Township (Pennsylvania); Chester County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Kalamazoo') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Kalamazoo (Michigan); Kalamazoo County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Zanesville') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Zanesville (Ohio); Muskingum County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Piqua') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Piqua (Ohio); Miami County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lyons') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lyons (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pioneer') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pioneer (Ohio); Williams County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Troy') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Troy (Ohio); Miami County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Louis') != -1:
                    if testvar.find('Missouri') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'St. Louis (Missouri)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Dayton') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Dayton (Ohio); Montgomery County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Miamisburg') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Miamisburg (Ohio); Montgomery County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Keokuk') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Keokuk (Iowa); Lee County (Iowa)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Washington Township') != -1 :
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Washington Township (Ohio); Lucas County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Orlando') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Orlando (Florida); Orange County (Florida)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Charleston') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Charleston (West Virginia); Kanawhad County (West Virginia)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bellefontaine') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bellefontaine (Ohio); Logan County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Berkey') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Berkey (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pittsfield') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pittsfield (Massachusetts); Berkshire County (Massachusetts)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Whitehouse') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Whitehouse (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Maumee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Maumee (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
        except:
            print('ERROR')
    iterationrow = iterationrow + 1
wb.save("aalh_iit_jeep_001.xlsx")