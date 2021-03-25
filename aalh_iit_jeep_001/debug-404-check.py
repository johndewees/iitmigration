from openpyxl import load_workbook
import urllib.request, urllib.error

filename = 'aalh_iit_jeep_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 43
maximumcol = 43
minimumrow = 7
maximumrow = 396

iterationrow = 7
targetcol = 48

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        #print(iterationrow)
        url = ws.cell(row=iterationrow, column=targetcol).value
        #print(url)
        if url == 'SKIP':
            continue
        else:
            try:
                conn = urllib.request.urlopen(url)
                #print('URL OPEN = SUCCESSFUL')
            except urllib.error.HTTPError as e:
                print(iterationrow,'|','HTTPError: {}'.format(e.code),'|',url)
            except urllib.error.URLError as e:
                print(iterationrow,'|','URLError: {}'.format(e.reason),'|',url)
    iterationrow = iterationrow + 1
print('********URL CHECK COMPLETE********')