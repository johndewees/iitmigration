from openpyxl import load_workbook
import re

filename = 'aalh_iit_samjones_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 37

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = None
            elif testvar.find('-') != -1:
                isovalue = testvar
                if isovalue.endswith('?'):
                    isovalue = isovalue[:-1]
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                continue
            elif testvar.find('/') != -1:
                continue
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            continue
    print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=isostandardcol).value)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_samjones_001.xlsx')