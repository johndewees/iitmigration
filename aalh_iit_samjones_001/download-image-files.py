from openpyxl import load_workbook
import urllib.request

filename = 'aalh_iit_samjones_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 50
minimumrow = 7
maximumrow = 37

targetcol = 48
iterationrow = 7

for row in ws.iter_rows(min_row=minimumrow, min_col=targetcol, max_row=maximumrow, max_col=targetcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        if testvar == 'SKIP':
            continue
        else:
            downloadurl = ws.cell(row=iterationrow, column=48).value
            downloadfilename = ws.cell(row=iterationrow, column=43).value
            download = urllib.request.urlretrieve(downloadurl, downloadfilename)
    iterationrow = iterationrow + 1
print('*****IMAGES DOWNLOADED*****')