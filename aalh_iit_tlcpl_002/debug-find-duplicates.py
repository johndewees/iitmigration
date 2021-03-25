from openpyxl import load_workbook

filename = 'aalh_iit_tlcpl_002.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 43
maximumcol = 43
minimumrow = 7
maximumrow = 143

iterationrow = 7
identifiercol = 25
filenamecol = 43

countfilename = dict()
countidentifier = dict()

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        #print(iterationrow)
        testvar1 = ws.cell(row=iterationrow, column=filenamecol).value
        #print(testvar1)
        if testvar1 not in countfilename:
            countfilename[testvar1] = 1
        else:
            countfilename[testvar1] = countfilename[testvar1] + 1
    for cell in row:
        testvar2 = ws.cell(row=iterationrow, column=identifiercol).value
        #print(testvar2)
        try:
            if testvar2 not in countfilename:
                countidentifier[testvar2] = 1
            else:
                countidentifier[testvar2] = countidentifier[testvar2] + 1
        except:
            continue
    iterationrow = iterationrow + 1
for file1 in countfilename:
    if countfilename[file1] > 1:
        print('Duplicate File Name:',file1, countfilename[file1])
for file2 in countidentifier:
    if countidentifier[file2] > 1:
        print('Duplicate Identifier:',file2, countidentifier[file2])
print('*Duplicate Check Completed*')
#print(countfilename)
#print(countidentifier)