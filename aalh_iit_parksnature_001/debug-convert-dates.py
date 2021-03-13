from openpyxl import load_workbook

filename = 'aalh_iit_parksnature_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 31
maximumcol = 31
minimumrow = 7
maximumrow = 493

iterationrow = 7
targetcol = 31

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        #print(testvar)
        if testvar == None:
            continue
            #print('No Date Digitized')
        elif testvar.find('/') != -1:
            testvarlist = testvar.split('/')
            testvaryear = testvarlist[2]
            testvaryear = testvaryear.strip()
            testvarmonth = testvarlist[0]
            testvarmonth = testvarmonth.strip()
            testvarmonth = int(testvarmonth)
            if testvarmonth < 10:
                testvarmonth = str(testvarmonth)
                testvarmonth = '0' + testvarmonth
            else:
                testvarmonth = str(testvarmonth)
            testvarday = testvarlist[1]
            testvarday = testvarday.strip()
            testvarday = int(testvarday)
            if testvarday < 10:
                testvarday = str(testvarday)
                testvarday = '0' + testvarday
            else:
                testvarday = str(testvarday)
            isodate = testvaryear + '-' + testvarmonth + '-' + testvarday
            ws.cell(row=iterationrow, column=targetcol).value = isodate
            print(iterationrow)
            print(ws.cell(row=iterationrow, column=targetcol).value)
        else:
            continue
            #print('Date is already formatted correctly')
    for cell in row:
        testvar2 = ws.cell(row=iterationrow, column=targetcol).value
        if testvar2 == None:
            continue
            #print('Still No Date Digitized')
        elif testvar2.find('-') != -1:
            length = len(testvar2)
            if length > 10:
                print('***CHECK THIS LINE FOR INCORRECT FORMATTING***')
            elif length < 10:
                print('***CHECK THIS LINE FOR INCORRECT FORMATTING***')
            #else:
                #print('Date is correctly formatted')
    iterationrow = iterationrow + 1
wb.save('aalh_iit_parksnature_001.xlsx')