from openpyxl import load_workbook

filename = 'aalh_iit_peopleportraits_008.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 550

iterationrow = 7
titlecol = 2

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=titlecol).value
    #print(testvar)
    for cell in row:
        if testvar.find(',') != -1:
            names = testvar.split(',')
            lastname = names[0]
            firstname = names[1]
            lastname = lastname.strip()
            firstname = firstname.strip()
            finaltitle = firstname + ' ' + lastname
            ws.cell(row=iterationrow, column=titlecol).value = finaltitle
            print(iterationrow)
            print(ws.cell(row=iterationrow, column=titlecol).value)
        else:
            continue
    iterationrow = iterationrow + 1
wb.save('aalh_iit_peopleportraits_008.xlsx')