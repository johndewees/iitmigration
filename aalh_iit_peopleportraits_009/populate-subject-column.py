from openpyxl import load_workbook

filename = 'aalh_iit_peopleportraits_009.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 507

iterationrow = 7
descol = 8
subcol = 9

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=subcol).value
    for cell in row:
        #print(iterationrow)
        if testvar == None:
            ws.cell(row=iterationrow, column=subcol).value = 'Persons. Photographs.'
            print(iterationrow, 'Persons')
    iterationrow = iterationrow + 1
print('*****COMPLETED*****')
wb.save('aalh_iit_peopleportraits_009.xlsx')