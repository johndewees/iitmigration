from openpyxl import load_workbook

filename = 'aalh_iit_peopleportraits_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 505

iterationrow = 7
descol = 8
subcol = 9

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=descol).value
    for cell in row:
        print(iterationrow)
        if testvar.find('house') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Dwellings. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('House') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Dwellings. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('dwelling') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Dwellings. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('Dwelling') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Dwellings. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('Church') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Churches. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('church') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Churches. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        iterationrow = iterationrow + 1
#wb.save('aalh_iit_peopleportraits_001.xlsx')