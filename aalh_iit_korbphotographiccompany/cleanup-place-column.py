from openpyxl import load_workbook

filename = 'aalh_iit_korbphotographiccompany.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 331

iterationrow = 7
desccol = 8
targetcol = 13

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=desccol).value
        if testvar == None:
            ws.cell(row=iterationrow, column=targetcol).value = ''
            print('Intentionally left blank')
        elif testvar.find('Philadelphia') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Philadelphia (Pennsylvania); Philadelphia County (Pennsylvania)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Cleveland') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Cleveland (Ohio); Cuyahoga County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Okolona') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Okolona (Ohio); Henry  County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Napoleon') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Napoleon (Ohio); Henry  County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Upper Sandusky') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Upper Sandusky (Ohio); Wyandot County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Sandusky') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Sandusky (Ohio); Erie County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Mt. Vernon') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Mt. Vernon (Ohio); Knox County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Detroit') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Detroit (Michigan); Wayne County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Sylvania') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Sylvania (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Oregon') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Oregon (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Brooklyn') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Brooklyn (Michigan); Jackson County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Knoxville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Knoxville (Tennessee); Knox County (Tennessee)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Ada, Ohio') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Ada (Ohio); Hardin County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Waterville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Waterville (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Bowling Green') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Bowling Green (Ohio); Wood County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Perrysburg') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Perrysburg (Ohio); Wood County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Ottokee') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Ottokee (Ohio); Fulton County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Marblehead') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Marblehead (Ohio); Ottawa County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Delta') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Delta (Ohio); Fulton County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Genoa') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Genoa (Ohio); Ottawa County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Grand Rapids') != -1:
            if testvar.find('Ohio') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Ohio); Wood County (Ohio)'
                print(ws.cell(row=iterationrow, column=targetcol).value)
            elif testvar.find('Michigan') != -1:
                ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Michigan); Kent County (Michigan)'
                print(ws.cell(row=iterationrow, column=targetcol).value)                
        elif testvar.find('Galloway') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Galloway (Ohio); Franklin County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Bryan') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Bryan (Ohio); Williams County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Pemberville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Pemberville (Ohio); Wood County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Bradner') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Bradner (Ohio); Wood County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Port Clinton') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Port Clinton (Ohio); Ottawa County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Monroeville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Monroeville (Ohio); Huron County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Dundee') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Dundee (Michigan); Monroe County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Fremont') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Fremont (Ohio); Sandusky County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Fayette') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Fayette (Ohio); Fulton County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Springfield') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Springfield (Ohio); Clark County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Ann Arbor') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Ann Arbor (Michigan); Washtenaw County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Tiffin') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Tiffin (Ohio); Seneca County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Delphos') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Delphos (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Sault St. Marie') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Sault St. Marie (Michigan), Chippewa County (Michigan)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Defiance') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Defiance (Ohio); Defiance County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Louisville') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Louisville (Kentucky); Jefferson County (Kentucky)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Put-in-Bay') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Put-in-Bay Township (Ohio); Ottawa County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('St. Paul') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Saint Paul') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Providence') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Providence Township (Ohio); Lucas County (Ohio)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Winona') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Winona (Minnesota); Winona County (Minnesota)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Galveston') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Galveston (Texas); Galveston County (Texas)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        elif testvar.find('Camden') != -1:
            ws.cell(row=iterationrow, column=targetcol).value = 'Camden (New Jersey); Camden County (New Jersey)'
            print(ws.cell(row=iterationrow, column=targetcol).value)
        else :
            print('No changes needed')
        iterationrow = iterationrow + 1
wb.save("aalh_iit_korbphotographiccompany.xlsx")