from openpyxl import load_workbook
import re

filename = 'aalh_iit_korbphotographiccompany.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 331

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = ''
                print('STATUS = NO DATE')
            elif testvar.find('1921-1923') != -1:
                ws.cell(row=iterationrow, column=isostandardcol).value = '1921; 1922; 1923'
            elif testvar.find('1896-1897') != -1:
                ws.cell(row=iterationrow, column=isostandardcol).value = '1896; 1897'
            elif testvar.find('-') != -1:
                isovalue = testvar
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                print('STATUS = AMERICAN DATE')
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            print('STATUS = PROBLEM')
        print(ws.cell(row=iterationrow, column=isostandardcol).value)
        iterationrow = iterationrow + 1
wb.save("aalh_iit_korbphotographiccompany.xlsx")