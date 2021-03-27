from openpyxl import load_workbook

filename = 'aalh_iit_buildings_011.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 47
maximumcol = 47
minimumrow = 7
maximumrow = 1357

iterationrow = 7
targetcol = 47
formatcol = 33
extentcol = 34

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        formatvar = ws.cell(row=iterationrow, column=targetcol).value
        print(formatvar)
        if formatvar == None:
            print('FIELD = NONE')
        elif formatvar.find('Negative') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Negatives'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('negative') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Negatives'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('Slide') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Slides'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('slide') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Slides'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('Post') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Postcard'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('post') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Postcard'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('Draw') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Drawings'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('draw') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Drawings'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('Black') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Black and white photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('black') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Black and white photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('b&w') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Black and white photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('B&W') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Black and white photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('Color') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Color photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        elif formatvar.find('color') != -1:
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture; Color photograph'
            print(ws.cell(row=iterationrow, column=formatcol).value)
        else :
            ws.cell(row=iterationrow, column=formatcol).value = 'Picture'
            print(ws.cell(row=iterationrow, column=formatcol).value)
    for cell in row:
        extentvar = ws.cell(row=iterationrow, column=targetcol).value
        if extentvar == None:
            print('FIELD = NONE')
        elif extentvar != None:
            extentvar1 = extentvar.split(';')
            for item in extentvar1:
                if item.find('Ex') != -1:
                    extentvar2 = item
                    extentvar3 = extentvar2.replace('Extent: ','')
                    extentvar4 = extentvar3.strip()
                    ws.cell(row=iterationrow, column=extentcol).value = extentvar4
                    print(extentvar4)
        else :
            print('EXTENT = CONFUSING')
    iterationrow = iterationrow + 1
wb.save('aalh_iit_buildings_011.xlsx')