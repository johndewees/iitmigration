from openpyxl import load_workbook
import re

filename = 'aalh_iit_buildings_011.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 1357

iterationrow = 7
targetcol = 15

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        cleandate = None
        approx = 'approximately '
        try:
            testvar2 = re.search('[a-z]\s19', testvar)
            if testvar == None:
                ws.cell(row=iterationrow, column=targetcol).value = ''
            elif testvar2:
                continue
            elif testvar.startswith('c'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('C'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('a'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.endswith('?'):
                cleandate = testvar[:-1]
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate
            elif testvar.find('-') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find(',') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find('/') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            else :
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = cleandate[0]
        except:
            print(iterationrow,'|',testvar,'|','STATUS = PROBLEM')
    print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=targetcol).value)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_buildings_011.xlsx')