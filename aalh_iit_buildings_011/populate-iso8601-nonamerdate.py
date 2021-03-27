from openpyxl import load_workbook
import re

filename = 'aalh_iit_buildings_011.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 1357

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = None
            testvar2 = re.search('[a-z]\s19', testvar)
            if testvar2:
                if testvar.find('Jan') != -1:
                    amermon = '01'
                elif testvar.find('jan') != -1:
                    amermon = '01'
                elif testvar.find('Feb') != -1:
                    amermon = '02'
                elif testvar.find('feb') != -1:
                    amermon = '02'
                elif testvar.find('Mar') != -1:
                    amermon = '03'
                elif testvar.find('mar') != -1:
                    amermon = '03'
                elif testvar.find('Apr') != -1:
                    amermon = '04'
                elif testvar.find('apr') != -1:
                    amermon = '04'
                elif testvar.find('May') != -1:
                    amermon = '05'
                elif testvar.find('may') != -1:
                    amermon = '05'
                elif testvar.find('Jun') != -1:
                    amermon = '06'
                elif testvar.find('jun') != -1:
                    amermon = '06'
                elif testvar.find('Jul') != -1:
                    amermon = '07'
                elif testvar.find('jul') != -1:
                    amermon = '07'
                elif testvar.find('Aug') != -1:
                    amermon = '08'
                elif testvar.find('aug') != -1:
                    amermon = '08'
                elif testvar.find('Sep') != -1:
                    amermon = '09'
                elif testvar.find('sep') != -1:
                    amermon = '09'
                elif testvar.find('Oct') != -1:
                    amermon = '10'
                elif testvar.find('oct') != -1:
                    amermon = '10'
                elif testvar.find('Nov') != -1:
                    amermon = '11'
                elif testvar.find('nov') != -1:
                    amermon = '11'
                elif testvar.find('Dec') != -1:
                    amermon = '12'
                elif testvar.find('dec') != -1:
                    amermon = '12'
                yearraw = testvar.split()
                yearraw1 = yearraw[1]
                yearraw2 = yearraw1.strip()
                isovalue2 = yearraw2 + '-' + amermon
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue2
            elif testvar.find('-') != -1:
                isovalue = testvar
                if isovalue.endswith('?'):
                    isovalue = isovalue[:-1]
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                continue
            elif testvar.find('/') != -1:
                continue
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            continue
    print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=isostandardcol).value)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_buildings_011.xlsx')