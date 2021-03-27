from openpyxl import load_workbook

filename = 'aalh_iit_buildings_011.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 1343

iterationrow = 7
titlecol = 2
covcol = 10

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=covcol).value
    for cell in row:
        if testvar.find('St. Clair') != -1:
            title = testvar.replace('St. Clair St.','St. Clair Street')
            ws.cell(row=iterationrow, column=covcol).value = title
        elif testvar.find('St.') != -1:
            title = testvar.replace('St.','Street')
            ws.cell(row=iterationrow, column=covcol).value = title
        elif testvar.find('Dr.') != -1:
            title = testvar.replace('Dr.','Drive')
            ws.cell(row=iterationrow, column=covcol).value = title
        elif testvar.find('Rd.') != -1:
            title = testvar.replace('Rd.','Road')
            ws.cell(row=iterationrow, column=covcol).value = title
        elif testvar.find('Ave.') != -1:
            title = testvar.replace('Ave.','Avenue')
            ws.cell(row=iterationrow, column=covcol).value = title
        elif testvar.find('Blvd.') != -1:
            title = testvar.replace('Blvd.','Boulevard')
            ws.cell(row=iterationrow, column=covcol).value = title
        else:
            continue
    print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=covcol).value)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_buildings_011.xlsx')