from openpyxl import load_workbook

filename = 'aalh_iit_buildings_04.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 529

iterationrow = 7
desccol = 8
targetcol = 13
rawcovcol = 49
covcol = 10
titlecol = 2

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    placefield = ws.cell(row=iterationrow, column=targetcol).value
    testvar = ws.cell(row=iterationrow, column=desccol).value
    for cell in row:
        print(iterationrow)
        try:
            if placefield != None: 
                print('FIELD ALREADY POPULATED')
            #if searching description field, comment out section below, will pull false positives from Maumee River
            elif placefield == None:
                if testvar.find('Maumee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Maumee (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Philadelphia') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Philadelphia (Pennsylvania); Philadelphia County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Elk Creek Township') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Elk Creek Township (Pennsylvania); Erie County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Cleveland') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Cleveland (Ohio); Cuyahoga County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Elmore') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Elmore (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bucyrus') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bucyrus (Ohio); Crawford County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Columbus') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Columbus (Ohio); Franklin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Okolona') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Okolona (Ohio); Henry  County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Holgate') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Holgate (Ohio); Henry  County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Napoleon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Napoleon (Ohio); Henry  County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Upper Sandusky') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Upper Sandusky (Ohio); Wyandot County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sandusky') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Sandusky (Ohio); Erie County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Castalia') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Castalia (Ohio); Erie County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Mt. Vernon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Mt. Vernon (Ohio); Knox County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Detroit') != -1:
                    if testvar.find('Pennsylvania') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Hometown (Pennsylvania); Schuylkill County (Pennsylvania)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Hometown') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Detroit (Michigan); Wayne County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Flint') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Flint (Michigan); Genesee County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sylvania') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Sylvania (Ohio); Lucas County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oregon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oregon (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Holland') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Holland (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Brooklyn') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Brooklyn (Michigan); Jackson County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Dowagaic') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Dowagiac (Michigan); Cass County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Knoxville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Knoxville (Tennessee); Knox County (Tennessee)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ada, Ohio') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ada (Ohio); Hardin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Waterville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Waterville (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Richfield Township') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Richfield Township (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bowling Green') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bowling Green (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Rossford') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Rossford (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Perrysburg') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Perrysburg (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Cincinnati') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Cincinnati (Ohio); Hamilton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Hamilton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Hamilton (Ohio); Butler County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ottokee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ottokee (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Marblehead') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Marblehead (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delta') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delta (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Genoa') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Genoa (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Grand Rapids') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Ohio); Wood County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grand Rapids (Michigan); Kent County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Niagara Falls') != -1:
                    if testvar.find('New York') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Niagara Falls (New York); Niagara County (New York)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Canada') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Niagara Falls (Canada)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)     
                elif testvar.find('Buffalo') != -1:
                    if testvar.find('New York') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Buffalo (New York); Erie County (New York)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delaware') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Delaware (Ohio); Delaware County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Erie') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Erie Township (Michigan); Monroe County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Pennsylvania') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Erie (Pennsylvania); Erie County (Pennsylvania)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Galloway') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Galloway (Ohio); Franklin County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bryan') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bryan (Ohio); Williams County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pemberville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pemberville (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bradner') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bradner (Ohio); Wood County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Port Clinton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Port Clinton (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Catawba') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Catawba Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Monroeville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Monroeville (Ohio); Huron County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Dundee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Dundee (Michigan); Monroe County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Ignance') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'St. Ignance (Michigan); Mackinac County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Ignance') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Mackinac Island (Michigan); Mackinac County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Monroe') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Monroe (Michigan); Monroe County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fremont') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fremont (Ohio); Sandusky County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fayette') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fayette (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Springfield') != -1:
                    if testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Springfield (Ohio); Clark County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Illinois') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Springfield (Illinois); Sangamon County (Illinois)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)    
                elif testvar.find('Ann Arbor') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ann Arbor (Michigan); Washtenaw County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Tiffin') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Tiffin (Ohio); Seneca County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delphos') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delphos (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sault St. Marie') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Sault St. Marie (Michigan), Chippewa County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Defiance') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Defiance (Ohio); Defiance County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Louisville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Louisville (Kentucky); Jefferson County (Kentucky)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Put-in-Bay') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Put-in-Bay Township (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('St. Paul') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Saint Paul') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Saint Paul (Minnesota); Ramsey County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Providence') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Providence Township (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Bono') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Bono (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Winona') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Winona (Minnesota); Winona County (Minnesota)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Galveston') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Galveston (Texas); Galveston County (Texas)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Camden') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Camden (New Jersey); Camden County (New Jersey)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Swanton') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Swanton (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Highland Park') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Highland Park (Michigan); Wayne County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Clay Center') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Clay Center (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Greenville') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Greenville (Ohio); Darke County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Lakeside') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Lakeside (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Wauseon') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Wauseon (Ohio); Fulton County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oak Harbor') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oak Harbor (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Stryker') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Stryker (Ohio); Williams County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Akron') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Akron (Ohio); Summit County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Johnsons Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Johnson's Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Sugar Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Sugar Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find("Johnson's Island") != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = "Johnson's Island (Ohio); Ottawa County (Ohio)"
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Riga') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Riga Township (Michigan); Lenawee County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Grandville') != -1:
                    if testvar.find('Michigan') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Grandville (Michigan); Kent County (Michigan)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Decatur') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Decatur (Illinois); Macon County (Illinois)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Adrian') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Adrian (Michigan); Lenawee County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Oberlin') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Oberlin (Ohio); Lorain County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Milwaukee') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Milwaukee (Wisconsin); Milwaukee County (Wisconsin)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Ballast Island') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Ballast Island (Ohio); Ottawa County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Marion') != -1:
                    if testvar.find('Indiana') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Marion (Indiana); Grant County (Indiana)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Ohio') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Marion (Ohio); Marion County (Ohio)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Fort Wayne') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Fort Wayne (Indiana); Allen County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Huntington') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Huntington (Indiana); Huntington County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Delphi') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Delphi (Indiana); Carroll County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Attica') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Attica (Indiana); Fountain County (Indiana)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Baltimore') != -1:
                    ws.cell(row=iterationrow, column=targetcol).value = 'Baltimore (Maryland)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Toledo') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Toledo (Ohio); Lucas County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Jacksonville') != -1 :
                    if testvar.find('Florida') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Jacksonville (Florida); Duval County (Florida)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                    elif testvar.find('Illinois') != -1:
                        ws.cell(row=iterationrow, column=targetcol).value = 'Jacksonville (Illinois); Morgan County (Illinois)'
                        print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Pittsburgh') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Pittsburgh (Pennsylvania); Allegheny County (Pennsylvania)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Kalamazoo') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Kalamazoo (Michigan); Kalamazoo County (Michigan)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
                elif testvar.find('Zanesville') != -1 :
                    ws.cell(row=iterationrow, column=targetcol).value = 'Zanesville (Ohio); Muskingum County (Ohio)'
                    print(ws.cell(row=iterationrow, column=targetcol).value)
        except:
            print('ERROR')
    iterationrow = iterationrow + 1
wb.save("aalh_iit_buildings_04.xlsx")