from openpyxl import load_workbook
import re

filename = 'aalh_iit_herrallongcollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 838

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = ''
            elif testvar.find('-') != -1:
                isovalue = testvar
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            print('STATUS = PROBLEM')
        print(ws.cell(row=iterationrow, column=isostandardcol).value)
        iterationrow = iterationrow + 1

wb.save("aalh_iit_herrallongcollection.xlsx")