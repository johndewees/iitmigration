from openpyxl import load_workbook
import re

filename = 'aalh_iit_peopleportraits_006.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 587

iterationrow = 7
covcol = 10
rawcovcol = 49

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=rawcovcol).value
    for cell in row:
        print(iterationrow)
        if testvar == None:
            print('Nothing to work with')
        elif testvar.find('1') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('2') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('3') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('4') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('5') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('6') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('7') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('8') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        elif testvar.find('9') != -1:
            address1 = testvar.split(',')
            address2 = address1[0]
            address3 = address2.strip()
            ws.cell(row=iterationrow, column=covcol).value = address3
            print(ws.cell(row=iterationrow, column=covcol).value)
        iterationrow = iterationrow + 1
wb.save("aalh_iit_peopleportraits_006.xlsx")