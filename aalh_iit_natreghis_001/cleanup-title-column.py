from openpyxl import load_workbook

filename = 'aalh_iit_natreghis_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 193

iterationrow = 7
titlecol = 2
rawcovcol = 49
covcol = 10

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=rawcovcol).value
    #print(testvar)
    for cell in row:
        if testvar.find(',') != -1:
            title1 = testvar.split(',')
            title2 = title1[0]
            title2 = title2.strip()
            ws.cell(row=iterationrow, column=covcol).value = title2
            if title2.find('St. Clair') != -1:
                title3 = title2.replace('St. Clair St.','St. Clair Street')
                ws.cell(row=iterationrow, column=covcol).value = title3
            elif title2.find('St.') != -1:
                title3 = title2.replace('St.','Street')
                ws.cell(row=iterationrow, column=covcol).value = title3
            elif title2.find('Dr.') != -1:
                title3 = title2.replace('Dr.','Drive')
                ws.cell(row=iterationrow, column=covcol).value = title3
            elif title2.find('Rd.') != -1:
                title3 = title2.replace('Rd.','Road')
                ws.cell(row=iterationrow, column=covcol).value = title3
            elif title2.find('Ave.') != -1:
                title3 = title2.replace('Ave.','Avenue')
                ws.cell(row=iterationrow, column=covcol).value = title3
            elif title2.find('Blvd.') != -1:
                title3 = title2.replace('Blvd.','Boulevard')
                ws.cell(row=iterationrow, column=covcol).value = title3
            else:
                ws.cell(row=iterationrow, column=covcol).value = title2
        else:
            continue
    print(iterationrow,'|',testvar,'|',ws.cell(row=iterationrow, column=covcol).value)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_natreghis_001.xlsx')