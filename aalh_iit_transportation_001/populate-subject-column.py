from openpyxl import load_workbook

filename = 'aalh_iit_transportation_001_uploaded.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 513

iterationrow = 7
descol = 8
subcol = 9

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=descol).value
    for cell in row:
        print(iterationrow)
        if testvar.find('Church') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Churches. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('church') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Churches. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('canal') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Canals. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('Canal') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Canals. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('Portrait') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Persons. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        elif testvar.find('portrait') != -1:
            ws.cell(row=iterationrow, column=subcol).value = 'Persons. Photographs.'
            print(ws.cell(row=iterationrow, column=subcol).value)
        iterationrow = iterationrow + 1
wb.save('aalh_iit_transportation_001_uploaded.xlsx')