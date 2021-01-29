from openpyxl import load_workbook

filename = 'aalh_iit_transportation_001_uploaded.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 524

iterationrow = 7
targetcol = 46
linkstring = 'Terms associated with the photograph are: '

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        iitdescription = ws.cell(row=iterationrow, column=minimumcol).value
        print(iitdescription)
        keywords = ws.cell(row=iterationrow, column=targetcol).value
        print(keywords)
        if iitdescription == None:
            descriptionmerged = linkstring + keywords
            descriptionfinal = descriptionmerged.replace("&#39;", "'")
            ws.cell(row=iterationrow, column=minimumcol).value = descriptionfinal
        else:
            descriptionmerged = iitdescription + ' ' + linkstring + keywords
            descriptionfinal = descriptionmerged.replace("&#39;", "'")
            ws.cell(row=iterationrow, column=minimumcol).value = descriptionfinal
        print(ws.cell(row=iterationrow, column=minimumcol).value)
        iterationrow = iterationrow + 1
wb.save('aalh_iit_transportation_001_uploaded.xlsx')