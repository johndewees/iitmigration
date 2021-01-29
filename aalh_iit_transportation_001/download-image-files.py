from openpyxl import load_workbook
import urllib.request

filename = 'aalh_iit_transportation_001_uploaded.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 50
minimumrow = 7
maximumrow = 513

targetcol = 48
iterationrow = 7

for row in ws.iter_rows(min_row=minimumrow, min_col=targetcol, max_row=maximumrow, max_col=targetcol):
    for cell in row:
            downloadurl = ws.cell(row=iterationrow, column=48).value
            downloadfilename = ws.cell(row=iterationrow, column=43).value
            download = urllib.request.urlretrieve(downloadurl, downloadfilename)
    iterationrow = iterationrow + 1
print('IMAGES DOWNLOADED')