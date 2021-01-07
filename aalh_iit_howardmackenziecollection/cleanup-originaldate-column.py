from openpyxl import load_workbook
import re

filename = 'aalh_iit_howardmackenziecollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 221

iterationrow = 7
targetcol = 15

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    for cell in row:
        print(iterationrow)
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        cleandate = None
        approx = 'approximately '
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=targetcol).value = ''
            elif testvar.find('1922? - 1925?') != -1:
                cleandate = 'approximately 1922 - 1925'
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find('1930-1931') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.endswith('?'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('c'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('C'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.startswith('a'):
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = approx + cleandate[0]
            elif testvar.find('-') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            elif testvar.find(',') != -1:
                cleandate = testvar
                ws.cell(row=iterationrow, column=targetcol).value = cleandate
            else :
                cleandate = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=targetcol).value = cleandate[0]
            print(ws.cell(row=iterationrow, column=targetcol).value)
        except:
            print('STATUS = PROBLEM')
        iterationrow = iterationrow + 1
#wb.save('aalh_iit_howardmackenziecollection.xlsx')