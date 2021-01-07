from openpyxl import load_workbook
import re

filename = 'aalh_iit_howardmackenziecollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 221

iterationrow = 7
targetcol = 2
locationcol = 13
datecol = 15

commaspace = ', '
space = ' '
openbracket = '['
closebrakcet = ']'

states = ['Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut','Delaware','Florida','Georgia','Hawaii','Idaho','Illinois','Indiana','Iowa','Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts','Michigan','Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada','New Hampshire','New Jersey','New Mexico','New York','North Carolina','North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island','South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont','Virginia','Washington','West Virginia','Wisconsin','Wyoming']

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        locationvar = ws.cell(row=iterationrow, column=locationcol).value
        if locationvar == None:
            locationvarfinal = ''
            #print(locationvarfinal)
        elif locationvar in states:
            locationvarfinal = commaspace + locationvar
            #print(locationvarfinal)
        else :
            locationvar1 = locationvar.split(';')
            locationvar2 = locationvar1[0]
            locationvar3 = locationvar2.split('(')
            locationvar4a = locationvar3[0]
            locationvar4b = locationvar3[1]
            locationvar5a = locationvar4a.strip()
            locationvar5b = locationvar4b[:-1]
            locationvarfinal = commaspace + locationvar5a + commaspace + locationvar5b
            #print(locationvarfinal)
    for cell in row:
        datevar = str(ws.cell(row=iterationrow, column=datecol).value)
        if datevar == None:
            datevarfinal = ' [date unknown]'
            #print(datevarfinal)
        elif datevar.startswith("approximately"):
            datevarfinal = space + openbracket + datevar + closebrakcet
            #print(datevarfinal)
        elif datevar.find('-') != -1:
            datevarregex = re.findall('\d\d\d\d', datevar)
            datevarfinal = commaspace + datevarregex[0]
            #print(datevarfinal)
        else:
            datevarfinal = commaspace + datevar
            #print(datevarfinal)
    for cell in row:
        title = str(ws.cell(row=iterationrow, column=targetcol).value)
        if title.find('1962') != -1:
            title = title[:-6]
        elif title.find('1963') != -1:
            title = title[:-6]
        elif title.find('1964') != -1:
            title = title[:-6]
        elif title.find('1967') != -1:
            title = title[:-6]
        elif title.find('1968') != -1:
            title = title[:-6]
        elif title.find('1971') != -1:
            title = title[:-6]
        elif title.find('1972') != -1:
            title = title[:-6]
        elif title.find('1973') != -1:
            title = title[:-6]
        elif title.find('1974') != -1:
            title = title[:-6]
        elif title.find('1975') != -1:
            title = title[:-6]
        elif title.find('1976') != -1:
            title = title[:-6]
        elif title.find('1977') != -1:
            title = title[:-6]
        elif title.find('1978') != -1:
            title = title[:-6]
        elif title.find('1979') != -1:
            title = title[:-6]
        elif title.find('1980') != -1:
            title = title[:-6]
        elif title.find('1981') != -1:
            title = title[:-6]
        elif title.find('1982') != -1:
            title = title[:-6]
        elif title.find('1983') != -1:
            title = title[:-6]
        titlefinal = title + locationvarfinal + datevarfinal
        ws.cell(row=iterationrow, column=targetcol).value = titlefinal
        print(titlefinal)
    iterationrow = iterationrow + 1
wb.save('aalh_iit_howardmackenziecollection.xlsx')