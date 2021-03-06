from openpyxl import load_workbook
import re

filename = 'aalh_iit_howardmackenziecollection.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 15
maximumcol = 15
minimumrow = 7
maximumrow = 221

iterationrow = 7
targetcol = 15
isostandardcol = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    print(iterationrow)
    for cell in row:
        testvar = ws.cell(row=iterationrow, column=targetcol).value
        print(testvar)
        isovalue = None
        try:
            if testvar == None:
                ws.cell(row=iterationrow, column=isostandardcol).value = ''
                print('STATUS = NO DATE')
            elif testvar.find('1922? - 1925?') != -1:
                isovalue = '1922; 1923; 1924; 1925'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('1930-1931') != -1:
                isovalue = '1930; 1931'
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find('-') != -1:
                isovalue = testvar
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue
            elif testvar.find(',') != -1:
                print('STATUS = AMERICAN DATE')
            else :
                isovalue = re.findall('\d\d\d\d', testvar)
                ws.cell(row=iterationrow, column=isostandardcol).value = isovalue[0]
        except:
            print('STATUS = PROBLEM')
        print(ws.cell(row=iterationrow, column=isostandardcol).value)
        iterationrow = iterationrow + 1
#wb.save('aalh_iit_howardmackenziecollection.xlsx')