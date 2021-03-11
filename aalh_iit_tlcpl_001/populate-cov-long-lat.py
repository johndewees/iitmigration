from openpyxl import load_workbook
import re

filename = 'aalh_iit_tlcpl_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 8
maximumcol = 8
minimumrow = 7
maximumrow = 268

iterationrow = 7
desccol = 8
covcol = 10
latcol = 11
longcol = 12

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=desccol).value
    for cell in row:
        print(iterationrow)
        if testvar == None:
            continue
        elif testvar.find('Old Main') != -1:
            continue
        elif testvar.find('Main Library') != -1:
            ws.cell(row=iterationrow, column=covcol).value = '325 Michigan Street'
            ws.cell(row=iterationrow, column=latcol).value = '41.654358'
            ws.cell(row=iterationrow, column=longcol).value = '-83.539662'
            print('MAIN LIBRARY')
        elif testvar.find('West Toledo') != -1:
            ws.cell(row=iterationrow, column=covcol).value = '1320 West Sylvania Avenue'
            ws.cell(row=iterationrow, column=latcol).value = '41.69297'
            ws.cell(row=iterationrow, column=longcol).value = '-83.572885'
            print('WEST TOLEDO')
    iterationrow = iterationrow + 1
wb.save("aalh_iit_tlcpl_001.xlsx")