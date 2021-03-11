from openpyxl import load_workbook
import re

filename = 'aalh_iit_tlcpl_001.xlsx'
wb = load_workbook(filename)
ws = wb['Metadata Template']

minimumcol = 2
maximumcol = 2
minimumrow = 7
maximumrow = 268

iterationrow = 7
titlecol = 2
desccol = 8
subcol = 9
placecol = 13
timeperiodcol = 14
dateoforigcol = 15
isodate = 16

for row in ws.iter_rows(min_row=minimumrow, min_col=minimumcol, max_row=maximumrow, max_col=maximumcol):
    testvar = ws.cell(row=iterationrow, column=titlecol).value
    testvar2 = ws.cell(row=iterationrow, column=desccol).value
    print(iterationrow)
    for cell in row:
        if testvar.find('Jones Junior High School') != -1:
            ws.cell(row=iterationrow, column=placecol).value = 'Toledo (Ohio); Lucas County (Ohio)'
            titlevar1 = testvar.split(',')
            titlevar2 = titlevar1[0]
            titlevar3 = titlevar2.strip()
            ws.cell(row=iterationrow, column=titlecol).value = titlevar3
            print('Title:')
            print(ws.cell(row=iterationrow, column=titlecol).value)
            ws.cell(row=iterationrow, column=subcol).value = 'Persons. Photographs.; Jones Junior High School (Toledo, Ohio)'
            print('Subject:')
            print(ws.cell(row=iterationrow, column=subcol).value)
            ws.cell(row=iterationrow, column=placecol).value = 'Toledo (Ohio); Lucas County (Ohio)'
            print('Place:')
            print(ws.cell(row=iterationrow, column=placecol).value)
            if testvar2.find('-19') != -1:
                datere = re.findall('\d\d\d\d-\d\d\d\d', testvar2)
                date1 = datere[0]
                date2 = date1.strip()
                date3 = date2.split('-')
                date4a = date3[0]
                date4b = date3[1]
                date5a = date4a.strip()
                date5b = date4b.strip()
                ws.cell(row=iterationrow, column=dateoforigcol).value = date2
                ws.cell(row=iterationrow, column=isodate).value = date5a + '; ' + date5b
                if testvar2.find('192') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1920s'
                elif testvar2.find('193') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1930s'
                elif testvar2.find('194') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1940s'
                elif testvar2.find('195') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1950s'
                elif testvar2.find('196') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1960s'
                elif testvar2.find('197') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1970s'
                elif testvar2.find('198') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1980s'
                elif testvar2.find('199') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '1990s'
                elif testvar2.find('200') != -1:
                    ws.cell(row=iterationrow, column=timeperiodcol).value = '2000s'
                print('Time Period:')
                print(ws.cell(row=iterationrow, column=timeperiodcol).value)
                print('Date of Original:')
                print(ws.cell(row=iterationrow, column=dateoforigcol).value)
                print('ISO 8601 Date:')
                print(ws.cell(row=iterationrow, column=isodate).value)
            elif testvar2.find('19') != -1:
                try:
                    datere = re.findall('\d\d\d\d-\d\d', testvar2)
                    date1 = datere[0]
                    date2 = date1.strip()
                    date3 = date2.split('-')
                    date4a = date3[0]
                    date4b = date3[1]
                    date5a = date4a.strip()
                    date5b = '19' + date4b.strip()
                    ws.cell(row=iterationrow, column=dateoforigcol).value = date5a + '-' + date5b
                    ws.cell(row=iterationrow, column=isodate).value = date5a + '; ' + date5b
                    if testvar2.find('192') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1920s'
                    elif testvar2.find('193') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1930s'
                    elif testvar2.find('194') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1940s'
                    elif testvar2.find('195') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1950s'
                    elif testvar2.find('196') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1960s'
                    elif testvar2.find('197') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1970s'
                    elif testvar2.find('198') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1980s'
                    elif testvar2.find('199') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1990s'
                    elif testvar2.find('200') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '2000s'
                except:
                    datere = re.findall('\d\d\d\d', testvar2)
                    date1 = datere[0]
                    date2 = date1.strip()
                    ws.cell(row=iterationrow, column=dateoforigcol).value = date2
                    ws.cell(row=iterationrow, column=isodate).value = date2
                    if testvar2.find('192') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1920s'
                    elif testvar2.find('193') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1930s'
                    elif testvar2.find('194') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1940s'
                    elif testvar2.find('195') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1950s'
                    elif testvar2.find('196') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1960s'
                    elif testvar2.find('197') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1970s'
                    elif testvar2.find('198') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1980s'
                    elif testvar2.find('199') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '1990s'
                    elif testvar2.find('200') != -1:
                        ws.cell(row=iterationrow, column=timeperiodcol).value = '2000s'
                print('Time Period:')
                print(ws.cell(row=iterationrow, column=timeperiodcol).value)
                print('Date of Original:')
                print(ws.cell(row=iterationrow, column=dateoforigcol).value)
                print('ISO 8601 Date:')
                print(ws.cell(row=iterationrow, column=isodate).value)
        else:
            continue
    iterationrow = iterationrow + 1
#wb.save("aalh_iit_tlcpl_001.xlsx")